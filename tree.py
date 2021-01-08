from anytree import Node, RenderTree
import openpyxl
import numpy as np

def get_tree_data(myfile):
    sheet = openpyxl.load_workbook(myfile).active
    size = [sheet.max_row,sheet.max_column]
    data = np.array([[Node(cell.value, depth=cell.column) for cell in row] for row in sheet.rows])

    for y in range(1,size[0]):
        for x in range(1,size[1]):
            if data[y,x].name:
                temp = data[:y,:x]
                for y0 in reversed(range(np.size(temp,0))):
                    for x0 in reversed(range(np.size(temp,1))):
                        if data[y0,x0].name:
                            data[y][x].parent = data[y0,x0]
                            break
                    if data[y0,x0].name:
                        break

    result = RenderTree(data[0,0]).by_attr('name')
    return(result)